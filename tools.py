"""
All tool functions that Claude can call.
All Excel structure knowledge comes from the inspector — nothing is hardcoded here.
"""

import openpyxl
import pandas as pd
from typing import Optional

from config import EXCEL_FILE_PATH, MONTHLY_BUDGET
from inspector import ExcelSchema, inspect_excel, normalize_month

# ── Module-level schema cache (loaded once per process) ──────────────────────

_SCHEMA: Optional[ExcelSchema] = None

MONTH_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}


def get_schema() -> ExcelSchema:
    """Return the cached schema, inspecting the file on first call."""
    global _SCHEMA
    if _SCHEMA is None:
        _SCHEMA = inspect_excel(EXCEL_FILE_PATH)
    return _SCHEMA


def clear_schema_cache() -> None:
    global _SCHEMA
    _SCHEMA = None


def _fmt(amount: float) -> str:
    return f"${amount:,.2f}"


def _month_key(month: str, year: int):
    num = MONTH_MAP.get(month.lower().strip())
    return (year, num) if num else None


# ── Data loaders ──────────────────────────────────────────────────────────────

def _load_df() -> pd.DataFrame:
    """
    Read the transaction sheet using the auto-detected schema.
    Returns a DataFrame with standardised column names and a 'month_key' column.
    """
    schema = get_schema()
    df = pd.read_excel(
        schema.file_path,
        sheet_name=schema.transaction_sheet,
        header=schema.header_row,
    )

    # Rename detected columns to standard internal names
    rename = {}
    if schema.columns.date:        rename[schema.columns.date]        = "Date"
    if schema.columns.month:       rename[schema.columns.month]       = "Month"
    if schema.columns.description: rename[schema.columns.description] = "Description"
    if schema.columns.category:    rename[schema.columns.category]    = "Category"
    if schema.columns.amount:      rename[schema.columns.amount]      = "Amount"
    if schema.columns.payment:     rename[schema.columns.payment]     = "Payment"
    if schema.columns.notes:       rename[schema.columns.notes]       = "Notes"
    df = df.rename(columns=rename)

    # Keep only the columns we mapped
    keep = [c for c in ["Date", "Month", "Description", "Category", "Amount", "Payment", "Notes"]
            if c in df.columns]
    df = df[keep].copy()

    # Clean amounts
    df = df.dropna(subset=["Amount"])
    df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce")
    df = df[df["Amount"] > 0].copy()

    # Build month_key from Month column if present, otherwise from Date
    if "Month" in df.columns:
        df["month_key"] = df["Month"].apply(normalize_month)
    elif "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["month_key"] = df["Date"].apply(
            lambda d: (d.year, d.month) if pd.notna(d) else None
        )
    else:
        df["month_key"] = None

    if "Category" in df.columns:
        df = df.dropna(subset=["Category"])

    df = df[df["month_key"].notna()].copy()
    return df


def _load_dashboard_data() -> Optional[dict]:
    """
    Read the dashboard sheet with data_only=True and return computed cell values.
    Returns {(year, month_int): {category: float, '_total': float}}, or None if
    no dashboard was detected.

    The dashboard is the authoritative source for monthly aggregates because some
    spreadsheets hardcode fixed costs (e.g. Rent) directly in dashboard cells
    rather than as individual transactions — reading only the transactions sheet
    would silently miss those amounts.
    """
    schema = get_schema()
    if not schema.has_dashboard or schema.dashboard is None:
        return None

    d = schema.dashboard
    wb = openpyxl.load_workbook(schema.file_path, data_only=True)
    ws = wb[d.sheet]

    result = {}
    for row_idx in range(d.data_start_row, d.data_end_row + 1):
        month_val = ws.cell(row=row_idx, column=d.month_col).value
        if month_val is None:
            continue
        mk = normalize_month(month_val)
        if mk is None:
            continue

        cats = {}
        for cat_name, col_idx in d.category_col_map.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            cats[cat_name] = float(val) if isinstance(val, (int, float)) else 0.0

        if d.total_col:
            total_val = ws.cell(row=row_idx, column=d.total_col).value
            cats["_total"] = float(total_val) if isinstance(total_val, (int, float)) \
                             else sum(cats.values())
        else:
            cats["_total"] = sum(cats.values())

        result[mk] = cats

    return result


def _category_data_for_month(month: str, year: int):
    """
    Return {category: float, '_total': float} for a month, or an error/message dict.
    Prefers dashboard data; falls back to aggregating from the transactions sheet.
    """
    mk = _month_key(month, year)
    if mk is None:
        return {"error": f"Unknown month '{month}'. Use a full name like 'March'."}

    dash = _load_dashboard_data()
    if dash and mk in dash:
        return dash[mk]                 # dashboard path (authoritative)

    # Fallback: compute from transactions
    df = _load_df()
    mdf = df[df["month_key"] == mk]
    if mdf.empty:
        return {"message": f"No data found for {month} {year}."}

    schema = get_schema()
    if "Category" not in mdf.columns:
        total = float(mdf["Amount"].sum())
        return {"_total": total, "_no_categories": True}

    by_cat = mdf.groupby("Category")["Amount"].sum()
    result = {cat: float(amt) for cat, amt in by_cat.items()}
    result["_total"] = float(by_cat.sum())
    return result


# ── Tool 1 — get_spending_by_category ────────────────────────────────────────

def get_spending_by_category(month: str, year: int) -> dict:
    """Total spending broken down by category for the given month."""
    try:
        data = _category_data_for_month(month, year)
        if "error" in data or "message" in data:
            return data

        cats = {k: v for k, v in data.items() if not k.startswith("_") and v > 0}
        cats_sorted = dict(sorted(cats.items(), key=lambda x: -x[1]))
        return {
            "month":       f"{month} {year}",
            "total":       _fmt(data["_total"]),
            "by_category": {cat: _fmt(amt) for cat, amt in cats_sorted.items()},
        }
    except Exception as e:
        return {"error": str(e)}


# ── Tool 2 — get_top_merchants ────────────────────────────────────────────────

def get_top_merchants(month: str, year: int, n: int = 5) -> dict:
    """Top N merchants/stores by total spend for the given month."""
    try:
        mk = _month_key(month, year)
        if mk is None:
            return {"error": f"Unknown month '{month}'."}

        df = _load_df()
        mdf = df[df["month_key"] == mk]
        if mdf.empty:
            return {"message": f"No transactions found for {month} {year}."}

        if "Description" not in mdf.columns:
            return {"message": "No merchant/description column detected in this file."}

        by_merchant = (
            mdf.groupby("Description")["Amount"]
            .sum()
            .sort_values(ascending=False)
            .head(n)
        )
        return {
            "month":         f"{month} {year}",
            "top_merchants": [
                {"merchant": m, "total_spent": _fmt(float(a))}
                for m, a in by_merchant.items()
            ],
        }
    except Exception as e:
        return {"error": str(e)}


# ── Tool 3 — get_monthly_summary ─────────────────────────────────────────────

def get_monthly_summary(month: str, year: int) -> dict:
    """Full monthly spending summary vs the configured monthly budget."""
    try:
        data = _category_data_for_month(month, year)
        if "error" in data or "message" in data:
            return data

        schema  = get_schema()
        budget  = schema.monthly_budget or MONTHLY_BUDGET
        total   = data["_total"]
        remaining   = budget - total
        utilization = (total / budget * 100) if budget else 0.0

        cats = {k: v for k, v in data.items() if not k.startswith("_") and v > 0}
        cats_sorted = dict(sorted(cats.items(), key=lambda x: -x[1]))

        return {
            "month":             f"{month} {year}",
            "total_expenses":    _fmt(total),
            "monthly_budget":    _fmt(budget),
            "remaining_budget":  _fmt(remaining),
            "budget_utilization": f"{utilization:.1f}%",
            "status":            "under budget" if remaining >= 0 else "over budget",
            "by_category":       {cat: _fmt(amt) for cat, amt in cats_sorted.items()},
        }
    except Exception as e:
        return {"error": str(e)}


# ── Tool 4 — compare_months ──────────────────────────────────────────────────

def compare_months(month1: str, year1: int, month2: str, year2: int) -> dict:
    """Category-by-category spending comparison between two months."""
    try:
        d1 = _category_data_for_month(month1, year1)
        d2 = _category_data_for_month(month2, year2)

        if "error" in d1: return d1
        if "error" in d2: return d2

        all_cats = sorted(
            k for k in set(d1) | set(d2)
            if not k.startswith("_")
        )

        comparison = {}
        changes    = []
        for cat in all_cats:
            a1     = d1.get(cat, 0.0)
            a2     = d2.get(cat, 0.0)
            if a1 == 0 and a2 == 0:
                continue
            change = a2 - a1
            pct    = ((change / a1) * 100) if a1 > 0 else (100.0 if a2 > 0 else 0.0)
            comparison[cat] = {
                f"{month1}_{year1}": _fmt(a1),
                f"{month2}_{year2}": _fmt(a2),
                "change":    _fmt(change),
                "pct_change": f"{pct:+.1f}%",
            }
            changes.append((cat, change))

        increases = sorted([(c, v) for c, v in changes if v > 0], key=lambda x: -x[1])
        decreases = sorted([(c, v) for c, v in changes if v < 0], key=lambda x: x[1])

        return {
            "comparison":              f"{month1} {year1} vs {month2} {year2}",
            f"{month1}_{year1}_total": _fmt(d1.get("_total", 0.0)),
            f"{month2}_{year2}_total": _fmt(d2.get("_total", 0.0)),
            "net_change":              _fmt(d2.get("_total", 0.0) - d1.get("_total", 0.0)),
            "by_category":             comparison,
            "biggest_increases":       [{"category": c, "increase": _fmt(v)} for c, v in increases[:3]],
            "biggest_decreases":       [{"category": c, "decrease": _fmt(v)} for c, v in decreases[:3]],
        }
    except Exception as e:
        return {"error": str(e)}


# ── Tool 5 — get_transactions ─────────────────────────────────────────────────

def get_transactions(
    month: str,
    year: int,
    category: Optional[str] = None,
    min_amount: Optional[float] = None,
) -> dict:
    """Filtered list of individual transactions for a given month."""
    try:
        mk = _month_key(month, year)
        if mk is None:
            return {"error": f"Unknown month '{month}'."}

        df = _load_df()
        mdf = df[df["month_key"] == mk]
        if mdf.empty:
            return {"message": f"No transactions found for {month} {year}."}

        if category and "Category" in mdf.columns:
            mdf = mdf[mdf["Category"].str.lower() == category.lower()]
        if min_amount is not None:
            mdf = mdf[mdf["Amount"] >= min_amount]
        if mdf.empty:
            return {"message": "No transactions matched the given filters."}

        txns = []
        for _, row in mdf.iterrows():
            d = row.get("Date")
            date_str = d.strftime("%Y-%m-%d") if pd.notna(d) and hasattr(d, "strftime") else "N/A"
            txns.append({
                "date":        date_str,
                "description": str(row["Description"]) if "Description" in row and pd.notna(row["Description"]) else "N/A",
                "category":    str(row["Category"])    if "Category"    in row and pd.notna(row["Category"])    else "N/A",
                "amount":      _fmt(row["Amount"]),
                "payment":     str(row["Payment"])     if "Payment"     in row and pd.notna(row["Payment"])     else "N/A",
            })

        txns.sort(key=lambda x: x["date"])
        return {
            "month":   f"{month} {year}",
            "filters": {
                "category":   category,
                "min_amount": _fmt(min_amount) if min_amount is not None else None,
            },
            "count":        len(txns),
            "total":        _fmt(sum(float(t["amount"].replace("$", "").replace(",", "")) for t in txns)),
            "transactions": txns,
        }
    except Exception as e:
        return {"error": str(e)}


# ── Tool definitions (JSON schemas) ──────────────────────────────────────────

def build_tool_definitions(categories: list) -> list:
    """Generate tool JSON schemas using the detected category list."""
    cat_str = ", ".join(categories) if categories else "auto-detected from file"
    return [
        {
            "name": "get_spending_by_category",
            "description": (
                f"Returns total spending broken down by category for a given month. "
                f"Available categories: {cat_str}. "
                f"Use this for questions like 'how much did I spend on groceries?' "
                f"or 'what were my biggest expense categories?'"
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "month": {"type": "string", "description": "Full month name, e.g. 'March'."},
                    "year":  {"type": "integer", "description": "4-digit year, e.g. 2026."},
                },
                "required": ["month", "year"],
            },
        },
        {
            "name": "get_top_merchants",
            "description": (
                "Returns the top N merchants/stores by total amount spent in a given month. "
                "Use this when the user asks where they spent the most money."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "month": {"type": "string", "description": "Full month name, e.g. 'March'."},
                    "year":  {"type": "integer", "description": "4-digit year, e.g. 2026."},
                    "n":     {"type": "integer", "description": "Number of top merchants to return. Defaults to 5."},
                },
                "required": ["month", "year"],
            },
        },
        {
            "name": "get_monthly_summary",
            "description": (
                "Returns a full monthly spending summary: total expenses, monthly budget, "
                "remaining budget, utilisation %, and breakdown by category. "
                "Use this for broad overview questions like 'how did I do in March?' "
                "or 'was I over budget?'"
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "month": {"type": "string", "description": "Full month name, e.g. 'March'."},
                    "year":  {"type": "integer", "description": "4-digit year, e.g. 2026."},
                },
                "required": ["month", "year"],
            },
        },
        {
            "name": "compare_months",
            "description": (
                "Compares spending between two months category by category, highlighting "
                "biggest increases and decreases. Use this for questions like "
                "'how did March compare to February?'"
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "month1": {"type": "string", "description": "First month name."},
                    "year1":  {"type": "integer", "description": "First month's year."},
                    "month2": {"type": "string", "description": "Second month name."},
                    "year2":  {"type": "integer", "description": "Second month's year."},
                },
                "required": ["month1", "year1", "month2", "year2"],
            },
        },
        {
            "name": "get_transactions",
            "description": (
                "Returns a filtered list of individual transactions for a given month. "
                "Optionally filter by category or minimum amount. "
                f"Valid categories: {cat_str}."
            ),
            "input_schema": {
                "type": "object",
                "properties": {
                    "month":      {"type": "string",  "description": "Full month name, e.g. 'March'."},
                    "year":       {"type": "integer", "description": "4-digit year, e.g. 2026."},
                    "category":   {"type": "string",  "description": "Optional category filter."},
                    "min_amount": {"type": "number",  "description": "Optional minimum transaction amount in USD."},
                },
                "required": ["month", "year"],
            },
        },
    ]


# ── Dispatcher ────────────────────────────────────────────────────────────────

_TOOL_FUNCTIONS = {
    "get_spending_by_category": get_spending_by_category,
    "get_top_merchants":        get_top_merchants,
    "get_monthly_summary":      get_monthly_summary,
    "compare_months":           compare_months,
    "get_transactions":         get_transactions,
}


def execute_tool(tool_name: str, tool_input: dict):
    func = _TOOL_FUNCTIONS.get(tool_name)
    if func is None:
        return {"error": f"Unknown tool: '{tool_name}'"}
    return func(**tool_input)
