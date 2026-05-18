"""
Auto-detects the structure of any expense Excel file.
Returns an ExcelSchema used by all tools — no hardcoded sheet names,
column positions, or category lists.
"""

import re
import openpyxl
import pandas as pd
from dataclasses import dataclass
from datetime import datetime
from typing import Optional, Dict, List, Tuple

# ── Keyword sets for column role detection ────────────────────────────────────
_KW: Dict[str, set] = {
    "date":        {"date", "transaction date", "trans date", "txn date", "dated"},
    "month":       {"month", "month/year", "period", "month name"},
    "description": {"description", "merchant", "payee", "vendor", "name", "store",
                    "item", "desc", "transaction", "details", "particulars"},
    "category":    {"category", "cat", "type", "expense type", "expense category",
                    "spending category", "tag"},
    "amount":      {"amount", "amount ($)", "amount(usd)", "cost", "price",
                    "debit", "credit", "spend", "spent", "value"},
    "payment":     {"payment", "payment method", "method", "card",
                    "paid with", "paid by", "mode", "payment mode"},
    "notes":       {"notes", "note", "memo", "remarks", "comment", "comments"},
}

_TX_SHEET_KW   = {"transactions", "transaction", "data", "expenses", "expense",
                   "records", "ledger", "history", "log", "sheet1"}
_DASH_SHEET_KW = {"dashboard", "summary", "overview", "report",
                   "analysis", "monthly", "breakdown"}

# Matches "Aug-25", "Mar-26", "January-2025", "Sep/26", etc.
_MONTH_RE = re.compile(
    r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[-/\s]\d{2,4}$",
    re.IGNORECASE,
)

# ── Shared date normalisation (imported by tools.py) ─────────────────────────

def normalize_month(val) -> Optional[Tuple[int, int]]:
    """Convert any Month-cell value to (year, month_int), or None if unparseable."""
    if val is None:
        return None
    if isinstance(val, (datetime, pd.Timestamp)):
        return (val.year, val.month)
    s = str(val).strip().rstrip("*").strip()
    for fmt in ["%b-%y", "%b-%Y", "%B-%Y", "%B-%y", "%Y-%m-%d", "%m/%Y"]:
        try:
            dt = datetime.strptime(s, fmt)
            return (dt.year, dt.month)
        except ValueError:
            pass
    return None


# ── Schema dataclasses ────────────────────────────────────────────────────────

@dataclass
class ColumnMap:
    date:        Optional[str] = None
    month:       Optional[str] = None
    description: Optional[str] = None
    category:    Optional[str] = None
    amount:      Optional[str] = None
    payment:     Optional[str] = None
    notes:       Optional[str] = None


@dataclass
class DashboardInfo:
    sheet:            str
    header_row:       int               # openpyxl 1-indexed
    data_start_row:   int               # openpyxl 1-indexed
    data_end_row:     int               # openpyxl 1-indexed
    month_col:        int               # openpyxl 1-indexed column index
    category_col_map: Dict[str, int]    # {category_name: openpyxl_col_idx}
    total_col:        Optional[int]     # openpyxl 1-indexed, or None


@dataclass
class ExcelSchema:
    file_path:         str
    transaction_sheet: str
    header_row:        int               # for pd.read_excel(header=X), 0-indexed
    columns:           ColumnMap
    categories:        List[str]
    monthly_budget:    Optional[float]
    date_range_start:  Optional[str]
    date_range_end:    Optional[str]
    has_dashboard:     bool
    dashboard:         Optional[DashboardInfo] = None


# ── Public entry point ────────────────────────────────────────────────────────

def inspect_excel(file_path: str) -> ExcelSchema:
    """
    Inspect any Excel expense file and return a fully-populated ExcelSchema.
    No assumptions are made about sheet names, column positions, or categories.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    tx_sheet                = _find_transaction_sheet(wb)
    header_row_idx, col_map = _detect_header_and_columns(wb[tx_sheet])
    categories, date_range  = _read_categories_and_range(
        file_path, tx_sheet, header_row_idx, col_map
    )
    dashboard = _find_dashboard(wb, categories)
    budget    = _detect_budget(wb, dashboard) if dashboard else None

    return ExcelSchema(
        file_path=file_path,
        transaction_sheet=tx_sheet,
        header_row=header_row_idx,
        columns=col_map,
        categories=categories,
        monthly_budget=budget,
        date_range_start=date_range[0] if date_range else None,
        date_range_end=date_range[1] if date_range else None,
        has_dashboard=dashboard is not None,
        dashboard=dashboard,
    )


# ── Step 1: find the transaction sheet ───────────────────────────────────────

def _find_transaction_sheet(wb) -> str:
    scores = {}
    for name in wb.sheetnames:
        s = name.lower().strip()
        name_score    = sum(2 for kw in _TX_SHEET_KW if kw in s)
        content_score = _score_sheet_columns(wb[name])
        scores[name]  = name_score + content_score
    return max(wb.sheetnames, key=lambda n: scores[n])


def _score_sheet_columns(ws) -> int:
    """Return how many column keyword roles the sheet's header rows cover."""
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        str_vals = [str(v).lower().strip() for v in row
                    if v is not None and isinstance(v, str)]
        if len(str_vals) < 2:
            continue
        score = sum(
            1 for kw_set in _KW.values()
            if any(_matches(v, kw_set) for v in str_vals)
        )
        if score >= 2:
            return score
    return 0


# ── Step 2: find header row + map columns ────────────────────────────────────

def _detect_header_and_columns(ws) -> Tuple[int, ColumnMap]:
    """
    Scan the first 20 rows for the one that best matches column keyword sets.
    Returns (0-indexed header row, ColumnMap).
    """
    best_idx   = 0
    best_score = -1

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        str_vals = [str(v).lower().strip() for v in row
                    if v is not None and isinstance(v, str)]
        if len(str_vals) < 2:
            continue
        score = sum(
            1 for kw_set in _KW.values()
            if any(_matches(v, kw_set) for v in str_vals)
        )
        # Bonus for the two most critical columns
        if any(_matches(v, _KW["amount"])   for v in str_vals): score += 3
        if any(_matches(v, _KW["category"]) for v in str_vals): score += 3
        if score > best_score:
            best_score = score
            best_idx   = row_idx

    # Build ColumnMap from the winning row
    header_row = list(
        ws.iter_rows(min_row=best_idx + 1, max_row=best_idx + 1, values_only=True)
    )[0]
    col_map  = ColumnMap()
    assigned = set()

    for cell_val in header_row:
        if cell_val is None:
            continue
        original = str(cell_val)
        lower    = original.lower().strip()
        clean    = re.sub(r"[^a-z\s]", "", lower).strip()

        best_role  = None
        best_match = 0
        for role, kw_set in _KW.items():
            if role in assigned:
                continue
            for kw in kw_set:
                s = 10 if (lower == kw or clean == kw) else \
                    7  if (lower.startswith(kw) or lower.endswith(kw)) else \
                    5  if (kw in lower) else 0
                if s > best_match:
                    best_match = s
                    best_role  = role

        if best_role and best_match > 0:
            setattr(col_map, best_role, original)   # keep the original column name
            assigned.add(best_role)

    return best_idx, col_map


# ── Step 3: read categories + date range ─────────────────────────────────────

def _read_categories_and_range(
    file_path: str, sheet: str, header_row: int, col_map: ColumnMap
) -> Tuple[List[str], Optional[Tuple[str, str]]]:
    df = pd.read_excel(file_path, sheet_name=sheet, header=header_row)

    categories: List[str] = []
    if col_map.category and col_map.category in df.columns:
        raw = df[col_map.category].dropna().unique()
        categories = sorted(str(c).strip() for c in raw if str(c).strip())

    date_range = None
    date_col = col_map.month or col_map.date
    if date_col and date_col in df.columns:
        keys = set()
        for val in df[date_col].dropna():
            mk = normalize_month(val)
            if mk:
                keys.add(mk)
        if keys:
            lo, hi = min(keys), max(keys)
            date_range = (
                datetime(lo[0], lo[1], 1).strftime("%b %Y"),
                datetime(hi[0], hi[1], 1).strftime("%b %Y"),
            )

    return categories, date_range


# ── Step 4: find dashboard ────────────────────────────────────────────────────

def _find_dashboard(wb, categories: List[str]) -> Optional[DashboardInfo]:
    candidates = [
        n for n in wb.sheetnames
        if any(kw in n.lower() for kw in _DASH_SHEET_KW)
    ]
    for sheet_name in candidates:
        info = _parse_monthly_table(wb[sheet_name])
        if info:
            return DashboardInfo(sheet=sheet_name, **info)
    return None


def _parse_monthly_table(ws) -> Optional[dict]:
    """
    Find a month×category grid in a worksheet.
    A valid grid has a column with ≥3 consecutive month-string values.
    Returns a dict of field values for DashboardInfo, or None.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    for col_idx in range(1, min(max_col + 1, 30)):
        month_rows = []
        for row_idx in range(1, min(max_row + 1, 60)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            is_month = (
                (isinstance(val, str) and _MONTH_RE.match(val.strip())) or
                (isinstance(val, datetime) and val.day == 1)
            )
            if is_month:
                month_rows.append(row_idx)

        if len(month_rows) < 3:
            continue

        # Header is the nearest non-empty row above the first month row
        header_row = month_rows[0] - 1
        while header_row > 0:
            row_vals = [ws.cell(row=header_row, column=c).value
                        for c in range(1, max_col + 1)]
            if any(v for v in row_vals if isinstance(v, str) and str(v).strip()):
                break
            header_row -= 1
        if header_row < 1:
            continue

        # Map header row cells → category names and TOTAL column.
        # Stop at TOTAL or the first None gap after collecting starts —
        # the dashboard may have unrelated tables to the right of the main one.
        cat_col_map: Dict[str, int] = {}
        total_col: Optional[int]   = None
        collecting = False

        for c in range(1, max_col + 1):
            h     = ws.cell(row=header_row, column=c).value
            label = str(h).strip() if h is not None else ""

            if not label:
                if collecting:
                    break       # first gap after data = end of this table
                continue

            collecting = True
            if c == col_idx:    # skip the month column itself
                continue

            label_lower = label.lower()
            if "total" in label_lower:
                total_col = c
                break           # everything right of TOTAL is a different table
            elif "budget" in label_lower or label_lower.startswith("vs"):
                continue        # skip delta/budget annotation columns
            else:
                cat_col_map[label] = c

        if not cat_col_map:
            continue

        return {
            "header_row":       header_row,
            "data_start_row":   month_rows[0],
            "data_end_row":     month_rows[-1],
            "month_col":        col_idx,
            "category_col_map": cat_col_map,
            "total_col":        total_col,
        }

    return None


# ── Step 5: detect budget ─────────────────────────────────────────────────────

def _detect_budget(wb, dashboard: DashboardInfo) -> Optional[float]:
    """
    Scan the dashboard sheet for a cell labelled "budget" and return the
    nearest adjacent numeric value.
    """
    ws = wb[dashboard.sheet]
    scan_rows = min(ws.max_row, dashboard.data_start_row - 1)

    for row_idx in range(1, scan_rows + 1):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and isinstance(val, str) and "budget" in val.lower():
                for dr in range(-1, 4):
                    neighbor = ws.cell(row=row_idx + dr, column=col_idx).value
                    if isinstance(neighbor, (int, float)) and 0 < neighbor < 1_000_000:
                        return float(neighbor)
    return None


# ── Shared helper ─────────────────────────────────────────────────────────────

def _matches(val: str, kw_set: set) -> bool:
    clean = re.sub(r"[^a-z\s]", "", val).strip()
    return val in kw_set or clean in kw_set or any(kw in val for kw in kw_set)
